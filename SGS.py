import datetime
import json
# import ijson
import os
import math
# import numpy as np
import socket
import util
import openpyxl
import random

import mbfl.mbfl_for
import mbfl.command
import sbfl.sbfl_for
import sbfl.command
from CHMBFL_Flow import Cluster_Fom
from mbfl.mutpolyn import mutation_trick
from mbfl.mutpolyn import new_mutation_trick
from data_codeflaws import MutationRule

usegetSuslist = True

# Rewrite = False
Rewrite = True

data_dirpath = './codeflaws_muti/mutinfo-new'


mbfl_for_list = [
    "mbfl.mbfl_for.Tarantula",
    "mbfl.mbfl_for.Op2",
    "mbfl.mbfl_for.Jaccard",
    "mbfl.mbfl_for.Ochiai",
    "mbfl.mbfl_for.Dstar",
    "mbfl.mbfl_for.GP13",
    "mbfl.mbfl_for.Naish1",
    "mbfl.mbfl_for.Barinel",
    "mbfl.mbfl_for.muse",
    "mbfl.mbfl_for.nsus",
]

sbfl_for_list = [
    "sbfl.sbfl_for.Tarantula",
    "sbfl.sbfl_for.Op2",
    "sbfl.sbfl_for.Jaccard",
    "sbfl.sbfl_for.Ochiai",
    "sbfl.sbfl_for.Dstar",
    "sbfl.sbfl_for.GP13",
    "sbfl.sbfl_for.Naish1",
    "sbfl.sbfl_for.Barinel",
]
class Reduction:
    def __init__(self, k=1):
        self.times = min(k, 1)

    def reSetList(self, obj):
        newList = []
        for value in obj:
            if value not in newList:
                newList.append(value)
        return newList

    def testanalysis(self, datas):
        data = datas['data']
        setMutantList = datas['setMutantList']

        # 获取所需数据
        or_list = data["or_list"]
        fom_list = data['fom_list']

        fom_list_simple = list(map(lambda x: sorted(x['message'],
                                                    key=lambda y: y[0]), fom_list))

        def DeltaT(or_list, out_list):
            Tpassm = sum(out_list)
            Tfailm = len(out_list)-sum(out_list)
            Tpassp = sum(or_list)
            Tfailp = len(or_list)-sum(or_list)
            DeltaTpassm = (Tpassm-Tpassp)/Tpassp
            DeltaTfailm = -1*(Tfailm-Tfailp)/Tfailp
            DeltaTm = DeltaTpassm+DeltaTfailm
            return DeltaTm

        mutantList_linedict = dict()
        for hom in setMutantList:
            if sum(list(map(lambda x: 1 if [x] in fom_list_simple else 0, hom['message']))) < 2:
                continue
            delta = sum(list(map(lambda x: DeltaT(or_list,
                                                  fom_list[fom_list_simple.index([x])]['out_list'])
                                 , hom['message'])))
            for fom_message in hom['message']:
                line = fom_message[0]
                if line not in mutantList_linedict:
                    mutantList_linedict[line] = []
                mutantList_linedict[line].append([delta, hom])

        setMutantList_new = []
        for line in mutantList_linedict.keys():
            for i in range(0, math.ceil(len(mutantList_linedict[line])*self.times)):
                hom = sorted(mutantList_linedict[line], key=lambda x: x[0], reverse=True)[i][1]
                if hom not in setMutantList_new:
                    setMutantList_new.append(hom)

        return setMutantList_new

    def samping(self, datas):
        setMutantList = datas['setMutantList']
        return random.sample(setMutantList, math.ceil(len(setMutantList)*self.times))

    def some(self, datas):
        setMutantList = datas['setMutantList']

        mutantList_linedict = dict()
        for hom in setMutantList:
            for fom_message in hom['message']:
                line = fom_message[0]
                if line not in mutantList_linedict:
                    mutantList_linedict[line] = []
                mutantList_linedict[line].append(hom)

        setMutantList_new = []
        for line in mutantList_linedict.keys():
            setMutantList_new += random.sample(mutantList_linedict[line],
                                               math.ceil(self.times*len(mutantList_linedict[line])))

        return self.reSetList(setMutantList_new)

    def wsome(self, datas):
        data = datas['data']
        setMutantList = datas['setMutantList']

        # 获取所需数据
        or_list = data["or_list"]
        sbfl_data = data["sbfl"]

        mutantList_linedict = dict()
        for hom in setMutantList:
            for fom_message in hom['message']:
                line = fom_message[0]
                if line not in mutantList_linedict:
                    mutantList_linedict[line] = []
                mutantList_linedict[line].append(hom)

        touple = sbfl.command.touple(or_list, sbfl_data['cov'])
        sus_list = dict()
        sumpower = 0
        for line, touple_list in touple.items():
            if line not in mutantList_linedict:
                continue
            sus = max(sbfl.sbfl_for.Dstar(touple_list), 0)
            sus_list[line] = sus
            sumpower += sus

        linepower = dict()
        for line in sus_list:
            linepower[line] = sus_list[line]/sumpower*len(sus_list)

        setMutantList_new = []
        for line in mutantList_linedict.keys():
            if line not in linepower:
                continue

            setMutantList_new += random.sample(mutantList_linedict[line],
                                               min(math.ceil(self.times*linepower[line]*len(mutantList_linedict[line])),
                                                   len(mutantList_linedict[line])))

        return self.reSetList(setMutantList_new)

    def fomanalysis(self, datas):

        data = datas['data']
        setMutantList = datas['setMutantList']
        formula = datas['formula']

        # 获取所需数据
        or_list = data["or_list"]
        fom_list = data['fom_list']

        fom_list_simple = list(map(lambda x: sorted(x['message'],
                                                    key=lambda y: y[0]), fom_list))

        def susH(formula, or_list, kill_list, out_list):
            f, p = 0, 0
            kf, kp, nf, np = 0, 0, 0, 0
            for j in range(len(or_list)):
                if or_list[j] == 1:
                    # 原始pass
                    if out_list[j] == 0:
                        # 变异体p -> f
                        f += 1
                else:
                    # 原始fail
                    if out_list[j] == 1:
                        # 变异体 f -> p
                        p += 1
                if or_list[j] == 1:
                    # 原始pass
                    if kill_list[j] == 1:
                        # 变异体n
                        np += 1
                    else:
                        # 变异体k
                        kp += 1
                else:
                    # 原始fail
                    if kill_list[j] == 1:
                        # 变异体n
                        kf += 1
                    else:
                        # 变异体k
                        nf += 1
            sus = eval(formula)([len(or_list)-sum(or_list), sum(or_list), [f], [p], f, p, [kf], [kp], [nf], [np]])
            return sus

        mutantList_linedict = dict()
        for hom in setMutantList:
            delta = sum(list(map(lambda x: susH(formula, or_list,
                                                fom_list[fom_list_simple.index([x])]['kill_list'],
                                                fom_list[fom_list_simple.index([x])]['out_list'])
                                 , hom['message'])))
            for fom_message in hom['message']:
                line = fom_message[0]
                if line not in mutantList_linedict:
                    mutantList_linedict[line] = []
                mutantList_linedict[line].append([delta, hom])

        print(sum(list(map(lambda line: len(mutantList_linedict[line]), mutantList_linedict.keys()))))

        setMutantList_new = []
        for line in mutantList_linedict.keys():
            for i in range(0, math.ceil(len(mutantList_linedict[line])*self.times)):
                hom = sorted(mutantList_linedict[line], key=lambda x: x[0], reverse=True)[i][1]
                if hom not in setMutantList_new:
                    setMutantList_new.append(hom)

        print(len(setMutantList_new))
        return setMutantList_new

    def null(self, datas):
        return datas['setMutantList']


    # 使用生成的变异体计算最终怀疑都
def getSuslistByMutantist(sus_data, data, setMutantList, reduceFunction):
    Fault_Record = data["Fault_Record"]
    or_list = data["or_list"]
    line_len = data['linelen']
    dataReturn = {}

    for reduction in reduceFunction:
        dataForm = sus_data.copy()
        reduceKey = str(reduction).split()[2].split('.')[1]


        if not 'fomanalysis' == reduceKey:
            newMutantList = reduction({'data': data, 'setMutantList': setMutantList})
            out_dic, kill_dic, precisions, varietys, homtime, homnum = Tools().readMutlist(newMutantList, Fault_Record)
            dataForm['totaltime'] += homtime
            dataForm['homnum'] += homnum
            touple = mbfl.command.GetTouleList(or_list, out_dic, kill_dic)
        for j, mbfl_for in enumerate(mbfl_for_list):
            mbfl.mbfl_for.type_mbfl = 'max'
            if 'fomanalysis' == reduceKey:
                newMutantList = reduction({'data': data, 'setMutantList': setMutantList, 'formula': mbfl_for})
                out_dic, kill_dic, precisions, varietys, homtime, homnum = Tools().readMutlist(newMutantList, Fault_Record)
                dataForm['totaltime'] += homtime
                dataForm['homnum'] += homnum
                touple = mbfl.command.GetTouleList(or_list, out_dic, kill_dic)
            for word in ['max', 'ave', 'frequency', 'none']:
                if word not in dataForm:
                    dataForm[word] = dict()
                dataForm[word][mbfl_for] = dict()
                mbfl.mbfl_for.type_mbfl = word
                sus_list = []
                for line in touple['linedata']:
                    touple_list = touple['tf'], \
                                  touple['tp'], \
                                  touple['linedata'][line]['f'], \
                                  touple['linedata'][line]['p'], \
                                  touple['f2p'], \
                                  touple['p2f'], \
                                  touple['linedata'][line]['kf'], \
                                  touple['linedata'][line]['kp'], \
                                  touple['linedata'][line]['nf'], \
                                  touple['linedata'][line]['np'],
                    sus = eval(mbfl_for)(touple_list)
                    sus_list.append([line, sus])

                sus_list_sort = sorted(sus_list, key=lambda x: x[1], reverse=True)
                ranks = util.Evaluation().rank(sus_list_sort, Fault_Record, line_len)

                rank = [[], [], []]
                for fault in ranks:
                    rank[0].append(fault[1])
                    rank[1].append(fault[2])
                    rank[2].append(fault[3])
                dataForm[word][mbfl_for]['rank_best'] = rank[0]
                dataForm[word][mbfl_for]['rank_average'] = rank[1]
                dataForm[word][mbfl_for]['rank_worst'] = rank[2]

                exam = [[], [], []]
                for ith in range(len(ranks)):
                    if line_len > 0:
                        exam[0].append(rank[0][ith]/line_len)
                        exam[1].append(rank[1][ith]/line_len)
                        exam[2].append(rank[2][ith]/line_len)
                    else:
                        exam[0].append(rank[0][ith])
                        exam[1].append(rank[1][ith])
                        exam[2].append(rank[2][ith])
                    if rank[0][ith] > rank[1][ith] or rank[1][ith] > rank[2][ith]:
                        print('exam异常', rank[0][ith], rank[1][ith],  rank[2][ith])

                dataForm[word][mbfl_for]['exam_best'] = exam[0]
                dataForm[word][mbfl_for]['exam_average'] = exam[1]
                dataForm[word][mbfl_for]['exam_worst'] = exam[2]

        dataForm['precision'] = precisions
        dataForm['variety'] = varietys
        if reduceKey == 'null':
            return dataForm
        dataReturn[reduceKey] = dataForm
    return dataReturn


class Get_sus:
    # mbfl指导生成hmbfl
    def get_sus_MBFLguide_hmbfl(self, data_json, times):
        # print('%s 读取数据' % datetime.datetime.now())
        if not data_json:
            return False
        try:
            doc = list(data_json.keys())[0]
            data = data_json[doc]
            del data_json


            # 获取所需数据
            or_list = data["or_list"]
            # or_out = data["or_out"]
            src_path = data['path']
            Fault_Record = data["Fault_Record"]
            fom_list = data['fom_list']
            hom_list_all = data['hom_list_all']
            # true_out = data["true_out"]
            sbfl_data = data["sbfl"]
            fomnum = len(fom_list)

            sus_data = dict()
            sus_data['Fault_Record'] = Fault_Record
            sus_data['totaltime'] = 0
        except Exception as e:
            print('数据缺失 %s' % e)
            return False

        try:
            path_list = src_path.split('/')
            while len(path_list) > 0:
                if not path_list[0] == 'access':
                    path_list.pop(0)
                else:
                    path_list.pop(0)
                    break
            # nowpath = os.path.abspath(os.path.join(os.path.dirname("__file__"),os.path.pardir))
            nowpath = os.getcwd()
            for part in path_list:
                nowpath = os.path.join(nowpath, part)
            line_len = len(util.File().read_line(nowpath))
        except:
            line_len = 0

        def get_som_out(som_message_self, hom_list_all_self):
            som_dict = {}
            for hom_all in hom_list_all_self:
                if hom_all["message"][0] == som_message_self[0] and hom_all["message"][1] == som_message_self[1]:
                    # if "out_or" not in hom_all:
                    #     continue
                    # if len(hom_all["out_or"]) == 0:
                    #     continue
                    for fom_message in som_message:
                        som_dict[fom_message[0]] = [hom_all["out_list"], hom_all["kill_list"]]
                    return som_dict
            return False

        # print('%s 计算一阶怀疑度' % datetime.datetime.now())
        # ------------------------------------------------
        # 计算一阶怀疑度
        mbfl.mbfl_for.type_mbfl = 'max'
        fom_out_dic = dict()
        fom_kill_dic = dict()
        for fom in fom_list:
            fom_message = fom["message"][0]
            fom_out_list = fom["out_list"]
            fom_kill_list = fom["kill_list"]
            sus_data['totaltime'] += fom["time"]
            if fom_message[0] not in fom_out_dic:
                fom_out_dic[fom_message[0]] = []
                fom_kill_dic[fom_message[0]] = []
            fom_out_dic[fom_message[0]].append(fom_out_list)
            fom_kill_dic[fom_message[0]].append(fom_kill_list)
        touple = mbfl.command.GetTouleList(or_list, fom_out_dic, fom_kill_dic)
        sus_list = []
        for line in touple['linedata']:
            touple_list = touple['tf'], \
                          touple['tp'], \
                          touple['linedata'][line]['f'], \
                          touple['linedata'][line]['p'], \
                          touple['f2p'], \
                          touple['p2f'], \
                          touple['linedata'][line]['kf'], \
                          touple['linedata'][line]['kp'], \
                          touple['linedata'][line]['nf'], \
                          touple['linedata'][line]['np'],
            sus = mbfl.mbfl_for.Ochiai(touple_list)
            sus_list.append([line, sus, -1, -1])

        # print('%s 获取组队数据' % datetime.datetime.now())

        # 获取组队数据
        line_pairs_list = Tools().guide_line_pair(sus_list, fomnum*times)
        fom_line_dict = dict()
        for fom in fom_list:
            line = fom['message'][0][0]
            if not line in fom_line_dict:
                fom_line_dict[line] = []
            fom_line_dict[line].append(fom)

        def pairable(pair_self, fom_line_dict_self):
            for line in pair_self:
                if line not in fom_line_dict_self:
                    return False
            return True

        hom_list_simple = list(map(lambda x: sorted(x['message'], key=lambda y: y[0]), hom_out_list))

        # print('%s 获取高阶元组' % datetime.datetime.now())

        # 获取元组输入字典

        sfterload_hom = []

        out_dic = dict()
        kill_dic = dict()
        som_num = 0
        # print('%s len %s ' % (datetime.datetime.now(), len(line_pairs_list)))
        for pair in line_pairs_list:
            som_pair_list = []
            if not pairable(pair, fom_line_dict):
                continue

            for fom0 in fom_line_dict[pair[0]]:
                for fom1 in fom_line_dict[pair[1]]:
                    som_pair_list.append([fom0['message'][0], fom1['message'][0]])

            random_som_pair_list = random.sample(som_pair_list, len(som_pair_list))
            while len(random_som_pair_list) > 0:
                som_message = random_som_pair_list.pop(0)
                try:
                    loc = hom_list_simple.index(som_message)
                except:
                    continue
                hom = hom_list_all[loc]
                sus_data['totaltime'] += hom['time']
                for fom_in_hom in hom['message']:
                    line = fom_in_hom[0]
                    if line not in out_dic:
                        out_dic[line] = []
                        kill_dic[line] = []
                    out_dic[line].append(hom['out_list'])
                    kill_dic[line].append(hom['kill_list'])
                som_num += 1
                break

            if len(random_som_pair_list) > 0:
                sfterload_hom += random_som_pair_list


            # for som_message in random.sample(som_pair_list, len(som_pair_list)):
            #     try:
            #         loc = hom_list_simple.index(som_message)
            #     except:
            #         continue
            #     hom = hom_list_all[loc]
            #     sus_data['totaltime'] += hom['time']
            #     for fom_in_hom in hom['message']:
            #         line = fom_in_hom[0]
            #         if line not in out_dic:
            #             out_dic[line] = []
            #             kill_dic[line] = []
            #         out_dic[line].append(hom['out_list'])
            #         kill_dic[line].append(hom['kill_list'])
            #     som_num += 1
            #     break
        # print('%s 补充高阶元组 %s- %s' % (datetime.datetime.now(), som_num , len(line_pairs_list)))
        while som_num < len(line_pairs_list) and len(sfterload_hom) > 0:
            som_message = sfterload_hom.pop(0)
            try:
                loc = hom_list_simple.index(som_message)
            except:
                continue
            hom = hom_list_all[loc]
            sus_data['totaltime'] += hom['time']
            for fom_in_hom in hom['message']:
                line = fom_in_hom[0]
                if line not in out_dic:
                    out_dic[line] = []
                    kill_dic[line] = []
                out_dic[line].append(hom['out_list'])
                kill_dic[line].append(hom['kill_list'])
            som_num += 1

        if som_num < len(line_pairs_list):
            print('变异体数量不足')
            return False

        # print('%s 计算高阶怀疑度' % datetime.datetime.now())

        for word in ['max', 'ave', 'frequency', 'none']:
            sus_data[word] = dict()
            mbfl.mbfl_for.type_mbfl = word
            touple = mbfl.command.GetTouleList(or_list, out_dic, kill_dic)
            for j, mbfl_for in enumerate(mbfl_for_list):
                sus_list = []
                sus_data[word][mbfl_for] = dict()
                for line in touple['linedata']:
                    touple_list = touple['tf'], \
                                  touple['tp'], \
                                  touple['linedata'][line]['f'], \
                                  touple['linedata'][line]['p'], \
                                  touple['f2p'], \
                                  touple['p2f'], \
                                  touple['linedata'][line]['kf'], \
                                  touple['linedata'][line]['kp'], \
                                  touple['linedata'][line]['nf'], \
                                  touple['linedata'][line]['np'],
                    sus = eval(mbfl_for)(touple_list)
                    sus_list.append([line, sus])


                sus_list_sort = sorted(sus_list, key=lambda x: x[1], reverse=True)
                ranks = util.Evaluation().rank(sus_list_sort, Fault_Record, line_len)

                rank = [[], [], []]
                for fault in ranks:
                    rank[0].append(fault[1])
                    rank[1].append(fault[2])
                    rank[2].append(fault[3])
                sus_data[word][mbfl_for]['rank_best'] = rank[0]
                sus_data[word][mbfl_for]['rank_average'] = rank[1]
                sus_data[word][mbfl_for]['rank_worst'] = rank[2]

                exam = [[], [], []]
                for ith in range(len(ranks)):
                    if line_len > 0:
                        exam[0].append(rank[0][ith]/line_len)
                        exam[1].append(rank[1][ith]/line_len)
                        exam[2].append(rank[2][ith]/line_len)
                    else:
                        exam[0].append(rank[0][ith])
                        exam[1].append(rank[1][ith])
                        exam[2].append(rank[2][ith])

                sus_data[word][mbfl_for]['exam_best'] = exam[0]
                sus_data[word][mbfl_for]['exam_average'] = exam[1]
                sus_data[word][mbfl_for]['exam_worst'] = exam[2]
        sus_data['totaltime'] += sbfl_data['time']
        return sus_data


def CHMBFL():

    def CHMBFL_mbfl_baseline_res():
        baseline_list = [
            ['Mbfl', Get_sus().Mbfl],
            # ['DeltaMbfl', Get_sus().DeltaMbfl],
            # ['Mcbfl', Get_sus().Mcbfl],
            # ['MUSE', Get_sus().Muse],
        ]
        ftitle = [
            'Id',
            'Version',
            'Fault_Record',
            'TimeSpend',
            'Fomnum',
            'Homnum',
            'Desrcibe',
        ]
        maxfaultnum = 3

        for method, function in baseline_list:
            # 表格初始化
            wb = openpyxl.Workbook()
            Wss = [wb.create_sheet('exam-best'), wb.create_sheet('exam-average'), wb.create_sheet('exam-worst')]
            del wb['Sheet']
            for ws in Wss:
                ws.cell(1, 1, method)
                for j, title in enumerate(ftitle):
                    ws.cell(2, j+1, title)
                for j, mbfl_for in enumerate(mbfl_for_list):
                    ws.cell(1, j*maxfaultnum*2+len(ftitle)+1, mbfl_for.split('.')[2])
                    for k in range(maxfaultnum):
                        ws.cell(2, j*maxfaultnum*2+len(ftitle)+1+k, 'exam%s' % str(k+1))
                        ws.cell(2, j*maxfaultnum*2+len(ftitle)+1+3+k, 'rank%s' % str(k+1))

            # 获取数据
            for file_i, file in enumerate(os.listdir(data_dirpath)):

                print('%s %s, read:%s-%s' % (datetime.datetime.now(), method, file, file_i))

                # 计算相应方法的怀疑度
                path = os.path.join(data_dirpath, file)
                sus_data_json = function(Tools().json_rules(path))

                sheetlist = [
                    [Wss[0], 'best'],
                    [Wss[1], 'average'],
                    [Wss[2], 'worst'],
                ]
                for sheet, name in sheetlist:
                    sheetdata_list = [
                        str(file_i+1),
                        file[5:-5],
                    ]
                    if not sus_data_json:
                        for sheetdata_i, sheetdata in enumerate(sheetdata_list):
                            sheet.cell(file_i + 3, sheetdata_i+1, sheetdata)
                        continue
                    sheetdata_list += [
                        len(sus_data_json['Fault_Record']),
                        sus_data_json['totaltime'],
                        sus_data_json['fomnum'],
                        sus_data_json['homnum'],
                        '',
                    ]
                    for mbfl_for_j, mbfl_for in enumerate(mbfl_for_list):
                        examlist = []
                        ranklist = []
                        for fault_ith in range(maxfaultnum):
                            if fault_ith > len(sus_data_json[mbfl_for]['rank_%s' % name])-1:
                                examlist.append('')
                                ranklist.append('')
                            else:
                                examlist.append(sus_data_json[mbfl_for]['exam_%s' % name][fault_ith])
                                ranklist.append(sus_data_json[mbfl_for]['rank_%s' % name][fault_ith])
                        sheetdata_list += examlist
                        sheetdata_list += ranklist
                    for sheetdata_i, sheetdata in enumerate(sheetdata_list):
                        sheet.cell(file_i + 3, sheetdata_i+1, sheetdata)

            docpath = os.path.join(os.getcwd(), 'report', 'CHMBFL', 'susfile', 'baseline', method, 'max')
            if not os.path.exists(docpath):
                os.makedirs(docpath)

            filepath = os.path.join(docpath, '%s_rank.xlsx' % method)
            wb.save(filepath)
            print('存储成功 %s' % filepath)

        return

    def CHMBFL_sbfl_baseline_res():
        # 初始化表格
        functionlist = [
            ['Sbfl', Get_sus().Sbfl],
        ]

        ftitle = [
            'Id',
            'Version',
            'Fault_Record',
            'TimeSpend',
            'Fomnum',
            'Homnum',
            'Desrcibe',
        ]
        maxfaultnum = 3

        for method, function in functionlist:
            # 表格初始化
            wb = openpyxl.Workbook()
            Wss = [wb.create_sheet('exam-best'), wb.create_sheet('exam-average'), wb.create_sheet('exam-worst')]
            del wb['Sheet']
            for ws in Wss:
                ws.cell(1, 1, method)
                for j, title in enumerate(ftitle):
                    ws.cell(2, j+1, title)
                for j, mbfl_for in enumerate(mbfl_for_list):
                    ws.cell(1, j*maxfaultnum*2+len(ftitle)+1, mbfl_for.split('.')[2])
                    for k in range(maxfaultnum):
                        ws.cell(2, j*maxfaultnum*2+len(ftitle)+1+k, 'exam%s' % str(k+1))
                        ws.cell(2, j*maxfaultnum*2+len(ftitle)+1+3+k, 'rank%s' % str(k+1))


            # 获取数据
            # sbfl_dirpath = './report/SBFL'
            for file_i, file in enumerate(os.listdir(data_dirpath)):
                print('%s %s, read:%s-%s' % (datetime.datetime.now(), method, file, file_i))
                # 计算相应方法的怀疑度
                path = os.path.join(data_dirpath, file)
                sus_data_json = function(Tools().json_rules(path))

                sheetlist = [
                    [Wss[0], 'best'],
                    [Wss[1], 'average'],
                    [Wss[2], 'worst'],
                ]
                for sheet, name in sheetlist:
                    sheetdata_list = [
                        str(file_i+1),
                        file[5:-5],
                    ]
                    if not sus_data_json:
                        for sheetdata_i, sheetdata in enumerate(sheetdata_list):
                            sheet.cell(file_i + 3, sheetdata_i+1, sheetdata)
                        continue
                    sheetdata_list += [
                        len(sus_data_json['Fault_Record']),
                        sus_data_json['totaltime'],
                        '',
                        '',
                        '',
                    ]
                    for sbfl_for_j, sbfl_for in enumerate(sbfl_for_list):
                        examlist = []
                        ranklist = []
                        for fault_ith in range(maxfaultnum):
                            if fault_ith > len(sus_data_json[sbfl_for]['rank_%s' % name])-1:
                                examlist.append('')
                                ranklist.append('')
                            else:
                                examlist.append(sus_data_json[sbfl_for]['exam_%s' % name][fault_ith])
                                ranklist.append(sus_data_json[sbfl_for]['rank_%s' % name][fault_ith])
                        sheetdata_list += examlist
                        sheetdata_list += ranklist
                    for sheetdata_i, sheetdata in enumerate(sheetdata_list):
                        sheet.cell(file_i + 3, sheetdata_i+1, sheetdata)



            # doc_rootpath = os.path.join(os.getcwd(), 'CHMBFL', 'susfile', 'baseline')
            doc_rootpath = os.path.join(os.getcwd(), 'report', 'CHMBFL', 'susfile', 'baseline', method)
            if not os.path.exists(doc_rootpath):
                os.makedirs(doc_rootpath)
            filepath = os.path.join(doc_rootpath, '%s_rank.xlsx' % method)
            wb.save(filepath)
            print('存储成功 %s' % filepath)

        return


    # 需要倍数， 需要重复次数
    def CHMBFL_timesneed_repeatneed_res():
        # 初始化表格
        clasterfunction = [
            ['Random', Get_sus().Random],
        ]
        ftitle = [
            'Id',
            'Version',
            'Fault_Record',
            'TimeSpend',
            'Fomnum',
            'Homnum',
            'Desrcibe',
        ]
        maxfaultnum = 3
        for method, function in clasterfunction:
            for times in range(1, 20):
                for repeat in range(1, 20):
                    # 表格初始化
                    Wbs = [openpyxl.Workbook(), openpyxl.Workbook(), openpyxl.Workbook()]
                    Wss = []
                    for wb in Wbs:
                        ws = [wb.create_sheet('exam-best'), wb.create_sheet('exam-average'), wb.create_sheet('exam-worst')]
                        del wb['Sheet']
                        for i in range(3):
                            ws[i].cell(1, 1, method)
                            for j, title in enumerate(ftitle):
                                ws[i].cell(2, j+1, title)
                            for j, mbfl_for in enumerate(mbfl_for_list):
                                ws[i].cell(1, j*maxfaultnum*2+len(ftitle)+1, mbfl_for.split('.')[2])
                                for k in range(maxfaultnum):
                                    ws[i].cell(2, j*maxfaultnum*2+len(ftitle)+1+k, 'exam%s' % str(k+1))
                                    ws[i].cell(2, j*maxfaultnum*2+len(ftitle)+1+3+k, 'rank%s' % str(k+1))
                        Wss.append(ws)

                    # 获取数据
                    for file_i, file in enumerate(os.listdir(data_dirpath)):
                        print('%s baseline:%s, times:%s repeat:%s read:%s-%s' %
                              (datetime.datetime.now(), method, times, repeat, file, file_i))
                        # 计算相应方法的怀疑度
                        path = os.path.join(data_dirpath, file)
                        sus_data_json = function(Tools().json_rules(path), times)

                        for ws_i, ws in enumerate(Wss):
                            word = ['max', 'ave', 'frequency'][ws_i]
                            sheetlist = [
                                [ws[0], 'best'],
                                [ws[1], 'average'],
                                [ws[2], 'worst'],
                            ]
                            for sheet, name in sheetlist:
                                sheetdata_list = [
                                    str(file_i+1),
                                    file[5:-5],
                                ]
                                if not sus_data_json:
                                    for sheetdata_i, sheetdata in enumerate(sheetdata_list):
                                        sheet.cell(file_i + 3, sheetdata_i+1, sheetdata)
                                    continue
                                sheetdata_list += [
                                    len(sus_data_json['Fault_Record']),
                                    sus_data_json['totaltime'],
                                    sus_data_json['fomnum'],
                                    sus_data_json['homnum'],
                                    '',
                                ]
                                for mbfl_for_j, mbfl_for in enumerate(mbfl_for_list):
                                    examlist = []
                                    ranklist = []
                                    for fault_ith in range(maxfaultnum):
                                        if fault_ith > len(sus_data_json[word][mbfl_for]['rank_%s' % name])-1:
                                            examlist.append('')
                                            ranklist.append('')
                                        else:
                                            examlist.append(sus_data_json[word][mbfl_for]['exam_%s' % name][fault_ith])
                                            ranklist.append(sus_data_json[word][mbfl_for]['rank_%s' % name][fault_ith])
                                    sheetdata_list += examlist
                                    sheetdata_list += ranklist
                                for sheetdata_i, sheetdata in enumerate(sheetdata_list):
                                    sheet.cell(file_i + 3, sheetdata_i+1, sheetdata)


                    # 存储文件
                    print('')
                    for wb_i, wb in enumerate(Wbs):
                        word = ['max', 'ave', 'frequency'][wb_i]
                        docpath = os.path.join(os.getcwd(), 'report/CHMBFL/susfile/numless', method, str(times), word)
                        if not os.path.exists(docpath):
                            os.makedirs(docpath)
                        filepath = os.path.join(docpath, '%s_%s_%s_%sth_ranke.xlsx' % (method, times, word, repeat))
                        wb.save(filepath)
                        print('文件保存 %s ' % filepath)
        return

    # 给定倍数，需要重复次数
    def CHMBFL_timesable_repeatneed_res():
        # 初始化表格
        clasterfunction = []
        if False:
            clasterfunction += [
                # ['DeltaNS', Get_sus().DeltaNS],
                ['DeltaNsMbfl', Get_sus().DeltaNsMbfl],
            ]
        if True:
            clasterfunction += [
                ['ED.MBFL', Get_sus().ErrorDistribution, [True]],
                ['ED.SBFL', Get_sus().ErrorDistribution, [False]],
                ['FTC', Get_sus().FailTestCluster, [True]],
                ['FTC.kmeans', Get_sus().FailTestCluster, [False]],
                ['MutCluster', Get_sus().MutCluster, [True, True]],
                ['MutCluster.in', Get_sus().MutCluster, [True, False]],
                ['MutCluster.kmeans', Get_sus().MutCluster, [False, True]],
                ['MutCluster.kmeans.in', Get_sus().MutCluster, [False, False]],
                ['NS.RANDOM', Get_sus().NS, [0]],
                ['NS.SBFL', Get_sus().NS, [1]],
                ['NS.MBFL', Get_sus().NS, [2]],
                ['NS.mbfl*FTC', Get_sus().NSpFTC],
                ['NS.mbfl*MC', Get_sus().NSpMC],
                ['NS.mbfl*ED', Get_sus().NSpED],
                ['MC.mseer*ED', Get_sus().MCpED],
                ['ED*NS*MC.mseer', Get_sus().EDpNSpMC],
                ['ED*NS*FTC.mseer', Get_sus().EDpNSpFTC],
            ]


        ftitle = [
            'Id',
            'Version',
            'Fault_Record',
            'TimeSpend',
            'Fomnum',
            'Homnum',
            'Desrcibe',

        ]
        # for i in range(13):
        #     ftitle.append('Operator Type %s' % str(i))
        # ftitle += [
        #     'None Accurate',
        #     'Pare Accurate',
        #     'Accurate',
        # ]
        maxfaultnum = 3
        # times = 11

        # for times in [1, 3, 5, 7, 10, 20, 50, 100]:
        for times in [0.5, 1, 1.5, 2, 2.5, 3, 3.5, 4, 4.5, 5]:
            for parameter in clasterfunction:
                method = parameter[0] + '-%s' % str(times)
                function = parameter[1]
                for repeat in range(1, 40):
                    # 表格初始化
                    Wbs = [openpyxl.Workbook(), openpyxl.Workbook(), openpyxl.Workbook(), openpyxl.Workbook()]
                    Wss = []
                    for wb in Wbs:
                        ws = [wb.create_sheet('exam-best'), wb.create_sheet('exam-average'), wb.create_sheet('exam-worst')]
                        del wb['Sheet']
                        for i in range(3):
                            ws[i].cell(1, 1, method)
                            for j, title in enumerate(ftitle):
                                ws[i].cell(2, j+1, title)
                            for j, mbfl_for in enumerate(mbfl_for_list):
                                ws[i].cell(1, j*maxfaultnum*2+len(ftitle)+1, mbfl_for.split('.')[2])
                                for k in range(maxfaultnum):
                                    ws[i].cell(2, j*maxfaultnum*2+len(ftitle)+1+k, 'exam%s' % str(k+1))
                                    ws[i].cell(2, j*maxfaultnum*2+len(ftitle)+1+3+k, 'rank%s' % str(k+1))
                        Wss.append(ws)

                    # 获取数据
                    for file_i, file in enumerate(os.listdir(data_dirpath)):

                        print('%s baseline:%s, repeat:%s read:%s-%s' %
                              (datetime.datetime.now(), method, repeat, file, file_i))
                        # 计算相应方法的怀疑度
                        path = os.path.join(data_dirpath, file)

                        if len(parameter) == 2:
                            sus_data_json = function(Tools().json_rules(path), times)
                        else:
                            sus_data_json = function(Tools().json_rules(path), parameter[2], times)

                        for ws_i, ws in enumerate(Wss):
                            word = ['max', 'ave', 'frequency', 'none'][ws_i]
                            sheetlist = [
                                [ws[0], 'best'],
                                [ws[1], 'average'],
                                [ws[2], 'worst'],
                            ]
                            for sheet, name in sheetlist:
                                sheetdata_list = [
                                    str(file_i+1),
                                    file[5:-5],
                                ]
                                if not sus_data_json:
                                    for sheetdata_i, sheetdata in enumerate(sheetdata_list):
                                        sheet.cell(file_i + 3, sheetdata_i+1, sheetdata)
                                    continue
                                sheetdata_list += [
                                    len(sus_data_json['Fault_Record']),
                                    sus_data_json['totaltime'],
                                    sus_data_json['fomnum'],
                                    sus_data_json['homnum'],
                                    '',
                                ]
                                # varietys = sus_data_json['variety']
                                # sum_varietys = sum(varietys)
                                # for variety in varietys:
                                #     try:
                                #         sheetdata_list.append(variety/sum_varietys)
                                #     except:
                                #         sheetdata_list.append(0)
                                #
                                # precisions = sus_data_json['precision']
                                # sum_precisions = sum(precisions)
                                # for precision in precisions:
                                #     try:
                                #         sheetdata_list.append(precision/sum_precisions)
                                #     except:
                                #         sheetdata_list.append(0)

                                for mbfl_for_j, mbfl_for in enumerate(mbfl_for_list):
                                    examlist = []
                                    ranklist = []
                                    for fault_ith in range(maxfaultnum):
                                        if fault_ith > len(sus_data_json[word][mbfl_for]['rank_%s' % name])-1:
                                            examlist.append('')
                                            ranklist.append('')
                                        else:
                                            examlist.append(sus_data_json[word][mbfl_for]['exam_%s' % name][fault_ith])
                                            ranklist.append(sus_data_json[word][mbfl_for]['rank_%s' % name][fault_ith])
                                    sheetdata_list += examlist
                                    sheetdata_list += ranklist
                                for sheetdata_i, sheetdata in enumerate(sheetdata_list):
                                    sheet.cell(file_i + 3, sheetdata_i+1, sheetdata)

                    # 存储文件
                    print('')
                    for wb_i, wb in enumerate(Wbs):
                        word = ['max', 'ave', 'frequency', 'none'][wb_i]
                        docpath = os.path.join(os.getcwd(), 'report/CHMBFL/susfile/K/', method, word)
                        if not os.path.exists(docpath):
                            os.makedirs(docpath)
                        filepath = os.path.join(docpath, '%s_%s_%sth_ranke.xlsx' % (method, word, repeat))
                        wb.save(filepath)
                        print('文件保存 %s ' % filepath)
        return

    # 给定倍数，需要重复次数
    def whf_res():
        # 初始化表格
        clasterfunction = []
        if True:
            clasterfunction += [
                # ['Last2First', Get_sus().Last2First],
                # ['DifferentOperator', Get_sus().DifferentOperator],
                # ['RandomMix', Get_sus().RandomMix],
                # ['Random', Get_sus().Random],

                ['ED.MBFL', Get_sus().ErrorDistribution, [True]],
                # ['ED.SBFL', Get_sus().ErrorDistribution, [False]],
                # ['FTC', Get_sus().FailTestCluster, [True]],
                # ['FTC.kmeans', Get_sus().FailTestCluster, [False]],
                # ['MutCluster', Get_sus().MutCluster, [True, True]],
                # ['MutCluster.in', Get_sus().MutCluster, [True, False]],
                # ['MutCluster.kmeans', Get_sus().MutCluster, [False, True]],
                # ['MutCluster.kmeans.in', Get_sus().MutCluster, [False, False]],
                # ['NS.RANDOM', Get_sus().NS, [0]],
                # ['NS.SBFL', Get_sus().NS, [1]],
                # ['NS.MBFL', Get_sus().NS, [2]],
                # ['NS.mbfl^FTC', Get_sus().NSpFTC],
                # ['NS.mbfl^MC', Get_sus().NSpMC],
                # ['NS.mbfl^ED', Get_sus().NSpED],
                # ['MC.mseer^ED', Get_sus().MCpED],
                # ['ED^NS^MC.mseer', Get_sus().EDpNSpMC],
                # ['ED^NS^FTC.mseer', Get_sus().EDpNSpFTC],
            ]
        ftitle = [
            'Id',
            'Version',
            'Fault_Record',
            'TimeSpend',
            'Fomnum',
            'Homnum',
            'Desrcibe',

        ]
        for i in range(15):
            ftitle.append('Operator Type %s' % str(i))
        ftitle += [
            'None Accurate',
            'Pare Accurate',
            'Accurate',
        ]
        maxfaultnum = 3

        try:
            # service = ['202.4.157.11', '202.4.157.19', '222.199.230.227', '202.4.130.29'].index(Tools().get_host_ip())
            service = ['202.4.130.29'].index(Tools().get_host_ip())
            timeslist = [
                [1, 2, 3, 4, 5]
            ][service]
            # timeslist = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11]
        except:
            # return 'IP 错误'
            timeslist = [11]

        for times in timeslist:
            for parameter in clasterfunction:
                method = parameter[0] + '-%s' % str(times)
                function = parameter[1]
                for repeat in range(1, 5):

                    if not Rewrite:
                        docpath = os.path.join(os.getcwd(), 'report/CHMBFL/susfile/nK', str(times), method, 'none')
                        filepath = os.path.join(docpath, '%s_%s_%sth_ranke.xlsx' % (method, 'none', repeat))
                        if os.path.exists(filepath):
                            continue


                    # 表格初始化
                    Wbs = [openpyxl.Workbook(), openpyxl.Workbook(), openpyxl.Workbook(), openpyxl.Workbook()]
                    Wss = []
                    for wb in Wbs:
                        ws = [wb.create_sheet('exam-best'), wb.create_sheet('exam-average'), wb.create_sheet('exam-worst')]
                        del wb['Sheet']
                        for i in range(3):
                            ws[i].cell(1, 1, method)
                            for j, title in enumerate(ftitle):
                                ws[i].cell(2, j+1, title)
                            for j, mbfl_for in enumerate(mbfl_for_list):
                                ws[i].cell(1, j*maxfaultnum*2+len(ftitle)+1, mbfl_for.split('.')[2])
                                for k in range(maxfaultnum):
                                    ws[i].cell(2, j*maxfaultnum*2+len(ftitle)+1+k, 'exam%s' % str(k+1))
                                    ws[i].cell(2, j*maxfaultnum*2+len(ftitle)+1+3+k, 'rank%s' % str(k+1))
                        Wss.append(ws)

                    # 获取数据
                    for file_i, file in enumerate(os.listdir(data_dirpath)):

                        print('%s baseline:%s, repeat:%s read:%s-%s' %
                              (datetime.datetime.now(), method, repeat, file, file_i))
                        # 计算相应方法的怀疑度
                        path = os.path.join(data_dirpath, file)

                        if len(parameter) == 2:
                            sus_data_json = function(Tools().json_rules(path), times=times)
                        else:
                            sus_data_json = function(Tools().json_rules(path), parameter[2], times)

                        for ws_i, ws in enumerate(Wss):
                            word = ['max', 'ave', 'frequency', 'none'][ws_i]
                            sheetlist = [
                                [ws[0], 'best'],
                                [ws[1], 'average'],
                                [ws[2], 'worst'],
                            ]
                            for sheet, name in sheetlist:
                                sheetdata_list = [
                                    str(file_i+1),
                                    file[5:-5],
                                ]
                                if not sus_data_json:
                                    for sheetdata_i, sheetdata in enumerate(sheetdata_list):
                                        sheet.cell(file_i + 3, sheetdata_i+1, sheetdata)
                                    continue
                                sheetdata_list += [
                                    len(sus_data_json['Fault_Record']),
                                    sus_data_json['totaltime'],
                                    sus_data_json['fomnum'],
                                    sus_data_json['homnum'],
                                    sus_data_json['Desrcibe'],
                                ]
                                varietys = sus_data_json['variety']
                                sum_varietys = sum(varietys)
                                for variety in varietys:
                                    try:
                                        sheetdata_list.append(variety/sum_varietys)
                                    except:
                                        sheetdata_list.append(0)

                                precisions = sus_data_json['precision']
                                sum_precisions = sum(precisions)
                                for precision in precisions:
                                    try:
                                        sheetdata_list.append(precision/sum_precisions)
                                    except:
                                        sheetdata_list.append(0)

                                for mbfl_for_j, mbfl_for in enumerate(mbfl_for_list):
                                    examlist = []
                                    ranklist = []
                                    for fault_ith in range(maxfaultnum):
                                        if fault_ith > len(sus_data_json[word][mbfl_for]['rank_%s' % name])-1:
                                            examlist.append('')
                                            ranklist.append('')
                                        else:
                                            examlist.append(sus_data_json[word][mbfl_for]['exam_%s' % name][fault_ith])
                                            ranklist.append(sus_data_json[word][mbfl_for]['rank_%s' % name][fault_ith])
                                    sheetdata_list += examlist
                                    sheetdata_list += ranklist
                                for sheetdata_i, sheetdata in enumerate(sheetdata_list):
                                    sheet.cell(file_i + 3, sheetdata_i+1, sheetdata)

                    # 存储文件
                    print('')
                    for wb_i, wb in enumerate(Wbs):
                        word = ['max', 'ave', 'frequency', 'none'][wb_i]
                        docpath = os.path.join(os.getcwd(), 'report/CHMBFL/susfile/nK', str(times), method, word)
                        if not os.path.exists(docpath):
                            os.makedirs(docpath)
                        filepath = os.path.join(docpath, '%s_%s_%sth_ranke.xlsx' % (method, word, repeat))
                        wb.save(filepath)
                        print('文件保存 %s ' % filepath)
        return

    # 给定倍数，需要重复次数
    def yojanTest():
        # 初始化表格
        reduction_list = [
            'testanalysis',
            'samping',
            'some',
            'wsome',
            'fomanalysis',
        ]
        clasterfunction = [
            ['DifferentOp', Get_sus().DifferentOperator],
            ['Last2First', Get_sus().Last2First],
            ['RandomMix', Get_sus().RandomMix],
            ['Random', Get_sus().Random],

            ['ErrorDistribution', Get_sus().ErrorDistribution, [False]],
            ['ErrorDistribution_SBFL', Get_sus().ErrorDistribution],

            ['MutCluster', Get_sus().MutCluster],
            ['MutCluster_kmeans', Get_sus().MutCluster, [False, True]],
            ['MutCluster_in', Get_sus().MutCluster, [True, False]],
            ['MutCluster_kmeans_in', Get_sus().MutCluster, [False, False]],

            ['NS_Random', Get_sus().NS],
            ['NS_MBFL', Get_sus().NS, [1]],
            ['NS_SBFL', Get_sus().NS, [2]],

            ['NS_MBFL^MC', Get_sus().NSpMC],
            ['NS_MBFL^ED_MBFL', Get_sus().NSpED],
            ['MC_mseer^ED_MBFL', Get_sus().MCpED],
            ['ED_MBFL^NS^MC_mseer', Get_sus().EDpNSpMC],
        ]
        ftitle = [
            'Id',
            'Version',
            'Fault_Record',
            'TimeSpend',
            'Fomnum',
            'Homnum',
            'Desrcibe',
        ]
        for i in range(15):
            ftitle.append('Operator Type %s' % str(i))
        ftitle += ['None Accurate', 'Pare Accurate', 'Accurate',]
        maxfaultnum = 3

        try:
            # service = ['202.4.157.11', '202.4.157.19', '222.199.230.227', '202.4.130.29'].index(Tools().get_host_ip())
            service = ['202.4.130.29'].index(Tools().get_host_ip())
            timeslist = [
                [0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1.0]
            ][service]
        except:
            # return 'IP 错误'
            timeslist = [0.1]

        for times in timeslist:
            for parameter in clasterfunction:
                for reductionname in reduction_list:
                    reductionfunction = eval('Reduction(%s).%s' % (str(times), reductionname))
                    method = '%s-%s-%s' % (reductionname, parameter[0], str(times).replace('.', '_'))
                    function = parameter[1]
                    for repeat in range(1, 6):
                        if not Rewrite:
                            docpath = os.path.join(os.getcwd(),
                                                   'report/CHMBFL/susfile/reduce',
                                                   str(times).replace('.', '_'),
                                                   method,
                                                   'max')
                            filepath = os.path.join(docpath, '%s_%s_%sth_ranke.xlsx' % (method, 'max', repeat))
                            if os.path.exists(filepath):
                                continue

                        # 表格初始化
                        Wbs = [openpyxl.Workbook(), openpyxl.Workbook(), openpyxl.Workbook(), openpyxl.Workbook()]
                        Wss = []
                        for wb in Wbs:
                            ws = [wb.create_sheet('exam-best'), wb.create_sheet('exam-average'), wb.create_sheet('exam-worst')]
                            del wb['Sheet']
                            for i in range(3):
                                ws[i].cell(1, 1, method)
                                for j, title in enumerate(ftitle):
                                    ws[i].cell(2, j+1, title)
                                for j, mbfl_for in enumerate(mbfl_for_list):
                                    ws[i].cell(1, j*maxfaultnum*2+len(ftitle)+1, mbfl_for.split('.')[2])
                                    for k in range(maxfaultnum):
                                        ws[i].cell(2, j*maxfaultnum*2+len(ftitle)+1+k, 'exam%s' % str(k+1))
                                        ws[i].cell(2, j*maxfaultnum*2+len(ftitle)+1+3+k, 'rank%s' % str(k+1))
                            Wss.append(ws)

                        # 获取数据
                        for file_i, file in enumerate(os.listdir(data_dirpath)):

                            print('%s baseline:%s, reduction:%s, repeat:%s read:%s-%s' %
                                  (datetime.datetime.now(), method, reductionname, repeat, file, file_i))
                            # if not '1562' in file:
                            #     continue
                            # 计算相应方法的怀疑度
                            path = os.path.join(data_dirpath, file)
                            data_json = Tools().json_rules(path)
                            # print(datetime.datetime.now())

                            # try:
                            if len(parameter) == 2:
                                sus_data_json = function(data_json, times=1, reducefunction=reductionfunction)
                            else:
                                sus_data_json = function(data_json, parameter[2], times=1, reducefunction=reductionfunction)
                            # except Exception as e:
                            #     print(e)
                            #     sus_data_json = False
                            # print(datetime.datetime.now())

                            for ws_i, ws in enumerate(Wss):
                                word = ['max', 'ave', 'frequency', 'none'][ws_i]
                                sheetlist = [
                                    [ws[0], 'best'],
                                    [ws[1], 'average'],
                                    [ws[2], 'worst'],
                                ]
                                for sheet, name in sheetlist:
                                    sheetdata_list = [
                                        str(file_i+1),
                                        file[5:-5],
                                    ]
                                    if not sus_data_json:
                                        for sheetdata_i, sheetdata in enumerate(sheetdata_list):
                                            sheet.cell(file_i + 3, sheetdata_i+1, sheetdata)
                                        continue
                                    sheetdata_list += [
                                        len(sus_data_json['Fault_Record']),
                                        sus_data_json['totaltime'],
                                        sus_data_json['fomnum'],
                                        sus_data_json['homnum'],
                                        sus_data_json['Desrcibe'],
                                    ]
                                    varietys = sus_data_json['variety']
                                    sum_varietys = sum(varietys)
                                    for variety in varietys:
                                        try:
                                            sheetdata_list.append(variety/sum_varietys)
                                        except:
                                            sheetdata_list.append(0)

                                    precisions = sus_data_json['precision']
                                    sum_precisions = sum(precisions)
                                    for precision in precisions:
                                        try:
                                            sheetdata_list.append(precision/sum_precisions)
                                        except:
                                            sheetdata_list.append(0)

                                    for mbfl_for_j, mbfl_for in enumerate(mbfl_for_list):
                                        examlist = []
                                        ranklist = []
                                        for fault_ith in range(maxfaultnum):
                                            if fault_ith > len(sus_data_json[word][mbfl_for]['rank_%s' % name])-1:
                                                examlist.append('')
                                                ranklist.append('')
                                            else:
                                                examlist.append(sus_data_json[word][mbfl_for]['exam_%s' % name][fault_ith])
                                                ranklist.append(sus_data_json[word][mbfl_for]['rank_%s' % name][fault_ith])
                                        sheetdata_list += examlist
                                        sheetdata_list += ranklist
                                    for sheetdata_i, sheetdata in enumerate(sheetdata_list):
                                        sheet.cell(file_i + 3, sheetdata_i+1, sheetdata)

                        # 存储文件
                        print('')
                        for wb_i, wb in enumerate(Wbs):
                            word = ['max', 'ave', 'frequency', 'none'][wb_i]
                            docpath = os.path.join(os.getcwd(), 'report/CHMBFL/susfile/reduce', str(times), method, word)
                            if not os.path.exists(docpath):
                                os.makedirs(docpath)
                            filepath = os.path.join(docpath, '%s_%s_%sth_ranke.xlsx' % (method, word, repeat))
                            wb.save(filepath)
                            print('文件保存 %s ' % filepath)
        return

    # 给定倍数，需要重复次数
    def testMessage():
        ftitle = [
            '版本',
            '代码行数',
            '测试用例个数',
            '错误数量',
            '生成一阶变异体数量',
        ]

        wb = openpyxl.Workbook()
        ws = wb['Sheet']
        for j, title in enumerate(ftitle):
            ws.cell(1, j+1, title)

        line = 2
        data_dirpath = './codeflaws_muti/mutinfo-new'
        for file_i, file in enumerate(os.listdir(data_dirpath)):
            print(file)
            path = os.path.join(data_dirpath, file)
            data_json = Tools().json_rules(path)
            doc = list(data_json.keys())[0]
            writeData = [
                doc,
                data_json[doc]['linelen'],
                len(data_json[doc]['or_list']),
                len(data_json[doc]['Fault_Record']),
                len(data_json[doc]['fom_list']),
            ]
            for j, value in enumerate(writeData):
                ws.cell(line, j+1, value)
            line += 1
        data_dirpath = './codeflaws_muti/mutinfo-single'
        for file_i, file in enumerate(os.listdir(data_dirpath)):
            print(file)
            path = os.path.join(data_dirpath, file)
            data_json = Tools().json_rules(path)
            doc = list(data_json.keys())[0]
            writeData = [
                doc,
                data_json[doc]['linelen'],
                len(data_json[doc]['or_list']),
                len(data_json[doc]['Fault_Record']),
                len(data_json[doc]['fom_list']),
            ]
            for j, value in enumerate(writeData):
                ws.cell(line, j+1, value)
            line += 1
        wb.save('codeflaws需要统计的数据集信息.xlsx')
        return


    # 给定倍数，需要重复次数
    def yojanTest_new():
        # 初始化表格
        clasterfunction = []
        reduction_list = [
            ['testanalysis', Reduction().testanalysis],
            ['samping', Reduction().samping],
            ['some', Reduction().some],
            ['wsome', Reduction().wsome],
            # ['none', Reduction().none],
        ]
        clasterfunction += [
            ['DifferentOp', Get_sus().DifferentOperator],
            ['Last2First', Get_sus().Last2First],
            ['RandomMix', Get_sus().RandomMix],
            ['Random', Get_sus().Random],
            ['ErrorDistribution', Get_sus().ErrorDistribution],
            ['MutCluster', Get_sus().MutCluster],
            ['NS_Random', Get_sus().NS],
            ['NS_SBFL', Get_sus().NS, [2]],
        ]
        ftitle = [
            'Id',
            'Version',
            'Fault_Record',
            'TimeSpend',
            'Fomnum',
            'Homnum',
            'Desrcibe',
        ]
        for i in range(13):
            ftitle.append('Operator Type %s' % str(i))
        ftitle += ['None Accurate', 'Pare Accurate', 'Accurate',]
        maxfaultnum = 3

        try:
            # service = ['202.4.157.11', '202.4.157.19', '222.199.230.227', '202.4.130.29'].index(Tools().get_host_ip())
            service = ['202.4.157.19', '222.199.230.148'].index(Tools().get_host_ip())
            timeslist = [
                [1],
                [10],
            ][service]
        except:
            # return 'IP 错误'
            timeslist = [5]

        for times in timeslist:
            for parameter in clasterfunction:
                for reductionname, reductionfunction in reduction_list:
                    method = '%s-%s-%s' % (reductionname, parameter[0], str(times))
                    function = parameter[1]
                    for repeat in range(1, 3):
                        if not Rewrite:
                            docpath = os.path.join(os.getcwd(), 'report/CHMBFL/susfile/K', str(times), method, 'none')
                            filepath = os.path.join(docpath, '%s_%s_%sth_ranke.xlsx' % (method, 'none', repeat))
                            if os.path.exists(filepath):
                                continue

                        # 表格初始化
                        Wbs = [openpyxl.Workbook(), openpyxl.Workbook(), openpyxl.Workbook(), openpyxl.Workbook()]
                        Wss = []
                        for wb in Wbs:
                            ws = [wb.create_sheet('exam-best'), wb.create_sheet('exam-average'), wb.create_sheet('exam-worst')]
                            del wb['Sheet']
                            for i in range(3):
                                ws[i].cell(1, 1, method)
                                for j, title in enumerate(ftitle):
                                    ws[i].cell(2, j+1, title)
                                for j, mbfl_for in enumerate(mbfl_for_list):
                                    ws[i].cell(1, j*maxfaultnum*2+len(ftitle)+1, mbfl_for.split('.')[2])
                                    for k in range(maxfaultnum):
                                        ws[i].cell(2, j*maxfaultnum*2+len(ftitle)+1+k, 'exam%s' % str(k+1))
                                        ws[i].cell(2, j*maxfaultnum*2+len(ftitle)+1+3+k, 'rank%s' % str(k+1))
                            Wss.append(ws)

                        # 获取数据
                        for file_i, file in enumerate(os.listdir(data_dirpath)):

                            print('%s baseline:%s, reduction:%s, repeat:%s read:%s-%s' %
                                  (datetime.datetime.now(), method, reductionname, repeat, file, file_i))
                            # if file_i <= 124:
                            #     continue
                            # 计算相应方法的怀疑度
                            path = os.path.join(data_dirpath, file)
                            data_json = Tools().json_rules(path)
                            reduction_datajson = reductionfunction(data_json, times/10)
                            # print(datetime.datetime.now())

                            try:
                                if len(parameter) == 2:
                                    sus_data_json = function(reduction_datajson, times=1000)
                                else:
                                    sus_data_json = function(reduction_datajson, parameter[2], times=1000)
                            except:
                                sus_data_json = False
                            # print(datetime.datetime.now())

                            for ws_i, ws in enumerate(Wss):
                                word = ['max', 'ave', 'frequency', 'none'][ws_i]
                                sheetlist = [
                                    [ws[0], 'best'],
                                    [ws[1], 'average'],
                                    [ws[2], 'worst'],
                                ]
                                for sheet, name in sheetlist:
                                    sheetdata_list = [
                                        str(file_i+1),
                                        file[5:-5],
                                    ]
                                    if not sus_data_json:
                                        for sheetdata_i, sheetdata in enumerate(sheetdata_list):
                                            sheet.cell(file_i + 3, sheetdata_i+1, sheetdata)
                                        continue
                                    sheetdata_list += [
                                        len(sus_data_json['Fault_Record']),
                                        sus_data_json['totaltime'],
                                        sus_data_json['fomnum'],
                                        sus_data_json['homnum'],
                                        sus_data_json['Desrcibe'],
                                    ]
                                    varietys = sus_data_json['variety']
                                    sum_varietys = sum(varietys)
                                    for variety in varietys:
                                        try:
                                            sheetdata_list.append(variety/sum_varietys)
                                        except:
                                            sheetdata_list.append(0)

                                    precisions = sus_data_json['precision']
                                    sum_precisions = sum(precisions)
                                    for precision in precisions:
                                        try:
                                            sheetdata_list.append(precision/sum_precisions)
                                        except:
                                            sheetdata_list.append(0)

                                    for mbfl_for_j, mbfl_for in enumerate(mbfl_for_list):
                                        examlist = []
                                        ranklist = []
                                        for fault_ith in range(maxfaultnum):
                                            if fault_ith > len(sus_data_json[word][mbfl_for]['rank_%s' % name])-1:
                                                examlist.append('')
                                                ranklist.append('')
                                            else:
                                                examlist.append(sus_data_json[word][mbfl_for]['exam_%s' % name][fault_ith])
                                                ranklist.append(sus_data_json[word][mbfl_for]['rank_%s' % name][fault_ith])
                                        sheetdata_list += examlist
                                        sheetdata_list += ranklist
                                    for sheetdata_i, sheetdata in enumerate(sheetdata_list):
                                        sheet.cell(file_i + 3, sheetdata_i+1, sheetdata)

                        # 存储文件
                        print('')
                        for wb_i, wb in enumerate(Wbs):
                            word = ['max', 'ave', 'frequency', 'none'][wb_i]
                            docpath = os.path.join(os.getcwd(), 'report/CHMBFL/susfile/K', str(times), method, word)
                            if not os.path.exists(docpath):
                                os.makedirs(docpath)
                            filepath = os.path.join(docpath, '%s_%s_%sth_ranke.xlsx' % (method, word, repeat))
                            wb.save(filepath)
                            print('文件保存 %s ' % filepath)
        return

    # 不需要倍数，需要重复次数
    def CHMBFL_baseline_timesable_repeatneed_res():
        clasterfunction = [
            ['Last2First', Get_sus().Last2First],
            ['DifferentOperator', Get_sus().DifferentOperator],
            ['RandomMix', Get_sus().RandomMix],
            ['Random', Get_sus().Random],
        ]
        ftitle = [
            'Id',
            'Version',
            'Fault_Record',
            'TimeSpend',
            'Fomnum',
            'Homnum',
            'Desrcibe',
        ]
        for i in range(15):
            ftitle.append('Operator Type %s' % str(i))
        ftitle += [
            'None Accurate',
            'Pare Accurate',
            'Accurate',
        ]
        maxfaultnum = 3

        try:
            # service = ['202.4.157.11', '202.4.157.19', '222.199.230.227', '202.4.130.29'].index(Tools().get_host_ip())
            service = ['202.4.157.19', '202.4.130.29'].index(Tools().get_host_ip())
            timeslist = [
                [1, 2, 3, 4, 5, 6],
                [7, 8, 9, 10, 11, 12, 13, 14, 15],
            ][service]
        except:
            # return 'IP 错误'
            timeslist = [0.5]

        for times in timeslist:
            # for method, function in clasterfunction:
            for parameter in clasterfunction:
                method = parameter[0] + '-%s' % str(times)
                function = parameter[1]
                for repeat in range(1, 5):
                    # 表格初始化
                    Wbs = [openpyxl.Workbook(), openpyxl.Workbook(), openpyxl.Workbook()]
                    Wss = []
                    for wb in Wbs:
                        ws = [wb.create_sheet('exam-best'), wb.create_sheet('exam-average'), wb.create_sheet('exam-worst')]
                        del wb['Sheet']
                        for i in range(3):
                            ws[i].cell(1, 1, method)
                            for j, title in enumerate(ftitle):
                                ws[i].cell(2, j+1, title)
                            for j, mbfl_for in enumerate(mbfl_for_list):
                                ws[i].cell(1, j*maxfaultnum*2+len(ftitle)+1, mbfl_for.split('.')[2])
                                for k in range(maxfaultnum):
                                    ws[i].cell(2, j*maxfaultnum*2+len(ftitle)+1+k, 'exam%s' % str(k+1))
                                    ws[i].cell(2, j*maxfaultnum*2+len(ftitle)+1+3+k, 'rank%s' % str(k+1))
                        Wss.append(ws)

                    # 获取数据
                    for file_i, file in enumerate(os.listdir(data_dirpath)):
                        print('%s baseline:%s, repeat:%s read:%s-%s' %
                              (datetime.datetime.now(), method, repeat, file, file_i))
                        # 计算相应方法的怀疑度
                        if not '1546' in file:
                            continue
                        path = os.path.join(data_dirpath, file)
                        sus_data_json = function(Tools().json_rules(path), times)

                        for ws_i, ws in enumerate(Wss):
                            word = ['max', 'ave', 'frequency', 'none'][ws_i]
                            sheetlist = [
                                [ws[0], 'best'],
                                [ws[1], 'average'],
                                [ws[2], 'worst'],
                            ]
                            for sheet, name in sheetlist:
                                sheetdata_list = [
                                    str(file_i+1),
                                    file[5:-5],
                                ]
                                if not sus_data_json:
                                    for sheetdata_i, sheetdata in enumerate(sheetdata_list):
                                        sheet.cell(file_i + 3, sheetdata_i+1, sheetdata)
                                    continue
                                sheetdata_list += [
                                    len(sus_data_json['Fault_Record']),
                                    sus_data_json['totaltime'],
                                    sus_data_json['fomnum'],
                                    sus_data_json['homnum'],
                                    sus_data_json['Desrcibe'],
                                ]
                                varietys = sus_data_json['variety']
                                sum_varietys = sum(varietys)
                                for variety in varietys:
                                    try:
                                        sheetdata_list.append(variety/sum_varietys)
                                    except:
                                        sheetdata_list.append(0)

                                precisions = sus_data_json['precision']
                                sum_precisions = sum(precisions)
                                for precision in precisions:
                                    try:
                                        sheetdata_list.append(precision/sum_precisions)
                                    except:
                                        sheetdata_list.append(0)

                                for mbfl_for_j, mbfl_for in enumerate(mbfl_for_list):
                                    examlist = []
                                    ranklist = []
                                    for fault_ith in range(maxfaultnum):
                                        if fault_ith > len(sus_data_json[word][mbfl_for]['rank_%s' % name])-1:
                                            examlist.append('')
                                            ranklist.append('')
                                        else:
                                            examlist.append(sus_data_json[word][mbfl_for]['exam_%s' % name][fault_ith])
                                            ranklist.append(sus_data_json[word][mbfl_for]['rank_%s' % name][fault_ith])
                                    sheetdata_list += examlist
                                    sheetdata_list += ranklist
                                for sheetdata_i, sheetdata in enumerate(sheetdata_list):
                                    sheet.cell(file_i + 3, sheetdata_i+1, sheetdata)

                    # 存储文件
                    print('')
                    for wb_i, wb in enumerate(Wbs):
                        word = ['max', 'ave', 'frequency', 'none'][wb_i]
                        # docpath = os.path.join(os.getcwd(), 'report/CHMBFL/susfile/baseline', method, word)
                        # docpath = os.path.join(os.getcwd(), 'report/CHMBFL/susfile/HomBaseline', str(times), method, word)
                        docpath = os.path.join(os.getcwd(), 'report/CHMBFL/susfile/K', str(times), method, word)
                        if not os.path.exists(docpath):
                            os.makedirs(docpath)
                        filepath = os.path.join(docpath, '%s_%s_%sth_ranke.xlsx' % (method, word, repeat))
                        wb.save(filepath)
                        print('文件保存 %s ' % filepath)
        return

    # 不需要倍数，需要重复次数
    def CHMBFL_timesless_repeatless_res():
        # 初始化表格
        clasterfunction = [
            # ['DifRank', Get_sus().test],
            ['RegroupBySus-New', Get_sus().NewRegroupBySus],
        ]
        ftitle = [
            'Id',
            'Version',
            'Fault_Record',
            'TimeSpend',
            'Fomnum',
            'Homnum',
            'Desrcibe',
        ]
        maxfaultnum = 3

        for method, function in clasterfunction:
            for repeat in range(1, 20):
                # 表格初始化
                Wbs = [openpyxl.Workbook(), openpyxl.Workbook(), openpyxl.Workbook()]
                Wss = []
                for wb in Wbs:
                    ws = [wb.create_sheet('exam-best'), wb.create_sheet('exam-average'), wb.create_sheet('exam-worst')]
                    del wb['Sheet']
                    for i in range(3):
                        ws[i].cell(1, 1, method)
                        for j, title in enumerate(ftitle):
                            ws[i].cell(2, j+1, title)
                        for j, mbfl_for in enumerate(mbfl_for_list):
                            ws[i].cell(1, j*maxfaultnum*2+len(ftitle)+1, mbfl_for.split('.')[2])
                            for k in range(maxfaultnum):
                                ws[i].cell(2, j*maxfaultnum*2+len(ftitle)+1+k, 'exam%s' % str(k+1))
                                ws[i].cell(2, j*maxfaultnum*2+len(ftitle)+1+3+k, 'rank%s' % str(k+1))
                    Wss.append(ws)


                # 获取数据
                for file_i, file in enumerate(os.listdir(data_dirpath)):
                    print('%s baseline:%s, repeat:%s read:%s-%s' %
                          (datetime.datetime.now(), method, repeat, file, file_i))
                    # 计算相应方法的怀疑度
                    path = os.path.join(data_dirpath, file)
                    sus_data_json = function(Tools().json_rules(path))

                    for ws_i, ws in enumerate(Wss):
                        word = ['max', 'ave', 'frequency'][ws_i]
                        sheetlist = [
                            [ws[0], 'best'],
                            [ws[1], 'average'],
                            [ws[2], 'worst'],
                        ]
                        for sheet, name in sheetlist:
                            sheetdata_list = [
                                str(file_i+1),
                                file[5:-5],
                            ]
                            if not sus_data_json:
                                for sheetdata_i, sheetdata in enumerate(sheetdata_list):
                                    sheet.cell(file_i + 3, sheetdata_i+1, sheetdata)
                                continue
                            sheetdata_list += [
                                len(sus_data_json['Fault_Record']),
                                sus_data_json['totaltime'],
                                sus_data_json['fomnum'],
                                sus_data_json['homnum'],
                                '',
                            ]
                            for mbfl_for_j, mbfl_for in enumerate(mbfl_for_list):
                                examlist = []
                                ranklist = []
                                for fault_ith in range(maxfaultnum):
                                    if fault_ith > len(sus_data_json[word][mbfl_for]['rank_%s' % name])-1:
                                        examlist.append('')
                                        ranklist.append('')
                                    else:
                                        examlist.append(sus_data_json[word][mbfl_for]['exam_%s' % name][fault_ith])
                                        ranklist.append(sus_data_json[word][mbfl_for]['rank_%s' % name][fault_ith])
                                sheetdata_list += examlist
                                sheetdata_list += ranklist
                            for sheetdata_i, sheetdata in enumerate(sheetdata_list):
                                sheet.cell(file_i + 3, sheetdata_i+1, sheetdata)

                # 存储文件
                print('')
                for wb_i, wb in enumerate(Wbs):
                    word = ['max', 'ave', 'frequency'][wb_i]
                    docpath = os.path.join(os.getcwd(), 'report/baseline', method, word)
                    if not os.path.exists(docpath):
                        os.makedirs(docpath)
                    filepath = os.path.join(docpath, '%s_%s_%sth_ranke.xlsx' % (method, word, repeat))
                    wb.save(filepath)
                    print('文件保存 %s ' % filepath)
        return

    # CHMBFL_baseline_timesable_repeatneed_res()
    # whf_res()
    testMessage()
    # CHMBFL_sbfl_baseline_res()
    # CHMBFL_mbfl_baseline_res()

    return

def newdata():
    # version_list = [
    #     '1553', '1571', '1591', '1597', '1623', '1627', '1647', '1666', '1690', '1708', '1721', '1724', '1735', '1751', '1752'
    # ]
    version_list = [
        # '1553',
        # '1571',
        '1751',
    ]
    for version in version_list:
        json_path = os.path.join('codeflaws/mutinfo', 'Fom_v%s_Feature.json' % version)
        with open(json_path) as f_obj:
            data_json = json.load(f_obj)
        print(version, 'mbfl')
        # print(version)
        data_mbfl = Get_sus().Mbfl(data_json)
        print(version, 'hmbfl')
        data_hmbfl = Get_sus().Random(data_json, 100000)


if __name__ == '__main__':
    CHMBFL()
    # newdata()
